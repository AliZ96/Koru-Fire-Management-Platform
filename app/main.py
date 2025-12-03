from fastapi import FastAPI, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse, FileResponse
from pathlib import Path
import asyncio, json, time, os, math
import re

# mevcut FIRMS try/except'in altına ekle:
from .weather import get_wind
from .spread import make_spread_sector, meters_to_deg

# Excel dosya okuma için openpyxl
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

# (opsiyonel) FIRMS; dosyan hazırsa çalışır, değilse ping yine çalışır
try:
    from .firms import fetch_firms_geojson
except Exception:
    fetch_firms_geojson = None

app = FastAPI(title="Izmir Wildfire Demo")

# Statik dosyaları /static altından servis et
app.mount("/static", StaticFiles(directory="static", html=True), name="static")

# Ana sayfa: login kontrolü ve yönlendirme
@app.get("/")
def home():
    return FileResponse("static/index.html")

# Login sayfası
@app.get("/login")
def login_page():
    return FileResponse("static/login.html")

# Welcome sayfası
@app.get("/welcome")
def welcome_page():
    return FileResponse("static/welcome.html")

# Basit sağlık kontrolü
@app.get("/api/ping")
def ping():
    return {"ok": True}

@app.get("/api/debug/env")
def debug_env():
    # Anahtarı göstermiyoruz; sadece var mı yok mu
    return {"has_map_key": bool(os.getenv("MAP_KEY"))}

@app.get("/api/fires_url")
def fires_url(day_range: int = 3):
    from app.firms import MAP_KEY, SOURCE, IZMIR_BBOX
    if not MAP_KEY:
        return {"error": "MAP_KEY not set"}
    url = f"https://firms.modaps.eosdis.nasa.gov/api/area/csv/{MAP_KEY}/{SOURCE}/{IZMIR_BBOX}/{day_range}"
    return {"test_url": url}
# FIRMS verisi (MAP_KEY gerekir)
@app.get("/api/fires")
def fires(day_range: int = 3):
    data = fetch_firms_geojson(day_range)
    if isinstance(data, dict) and "error" in data:
        from fastapi import status
        code = int(data.get("status", 500))
        return JSONResponse(data, status_code=code)
    return JSONResponse(data)
# 🔹 Rüzgâr: verilen noktada (varsayılan İzmir merkez) saatlik rüzgâr özeti
@app.get("/api/wind")
def wind(lat: float = Query(38.42), lon: float = Query(27.14), when: str | None = Query(None)):
    """
    lat, lon ve (opsiyonel) when='YYYY-MM-DDTHH:00' alır.
    Dönen: { speed_ms, deg, source, time }
    """
    w = get_wind(lat, lon, when_iso=when)
    return JSONResponse(w)

# 🔹 Yayılım: rüzgâr ve süreye göre sektör/oval poligon (GeoJSON)
@app.get("/api/spread")
def spread(
    lat: float = Query(38.42),
    lon: float = Query(27.14),
    wind_dir_deg: float = Query(240.0),
    wind_speed_ms: float = Query(6.0),
    duration_min: float = Query(30.0),
):
    feat = make_spread_sector(lat, lon, wind_dir_deg, wind_speed_ms, duration_min)
    return JSONResponse({"type": "FeatureCollection", "features": [feat]})

# 🔹 Toplanma Alanları: Excel dosyalarından oku ve GeoJSON olarak döndür
@app.get("/api/shelters_from_excel")
def shelters_from_excel():
    """
    static/data/toplanma-alanları/ klasöründeki Excel dosyalarını oku
    Beklenen kolonlar: ad/name, enlem/latitude/lat, boyam/longitude/lon
    """
    shelters_dir = Path("static/data/toplanma-alanları")
    features = []
    
    if not shelters_dir.exists():
        return {"type": "FeatureCollection", "features": [], "error": "Toplanma alanları klasörü bulunamadı"}
    
    # Excel dosyalarını bul
    excel_files = list(shelters_dir.glob("*.xlsx")) + list(shelters_dir.glob("*.xls"))
    
    for excel_file in excel_files:
        try:
            # openpyxl ile .xlsx okuyabilir
            if HAS_OPENPYXL and excel_file.suffix.lower() == '.xlsx':
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                
                # İlk satır header'ı varsayalım
                headers = []
                for cell in ws[1]:
                    headers.append((cell.value or "").lower().strip())
                
                # Enlem, boyam, ad sütunlarını bul
                lat_col = lon_col = name_col = None
                for i, h in enumerate(headers):
                    if 'enlem' in h or 'latitude' in h or 'lat' in h:
                        lat_col = i
                    elif 'boyam' in h or 'longitude' in h or 'lon' in h:
                        lon_col = i
                    elif 'ad' in h or 'name' in h:
                        name_col = i
                
                # Satırları oku
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or all(v is None for v in row):
                        continue
                    
                    try:
                        lat = float(row[lat_col]) if lat_col is not None and row[lat_col] else None
                        lon = float(row[lon_col]) if lon_col is not None and row[lon_col] else None
                        name = str(row[name_col]) if name_col is not None and row[name_col] else f"Toplanma Alanı ({excel_file.stem})"
                        
                        if lat and lon:
                            features.append({
                                "type": "Feature",
                                "geometry": {"type": "Point", "coordinates": [lon, lat]},
                                "properties": {"name": name, "ilçe": excel_file.stem.split("_")[0]}
                            })
                    except (ValueError, TypeError, IndexError):
                        continue
            
            # xlrd ile .xls okuyabilir
            elif HAS_XLRD and excel_file.suffix.lower() == '.xls':
                wb = xlrd.open_workbook(str(excel_file))
                ws = wb.sheet_by_index(0)
                
                # İlk satır header
                headers = []
                for cell in ws.row(0):
                    headers.append((cell.value or "").lower().strip())
                
                lat_col = lon_col = name_col = None
                for i, h in enumerate(headers):
                    if 'enlem' in h or 'latitude' in h or 'lat' in h:
                        lat_col = i
                    elif 'boyam' in h or 'longitude' in h or 'lon' in h:
                        lon_col = i
                    elif 'ad' in h or 'name' in h:
                        name_col = i
                
                for row_idx in range(1, ws.nrows):
                    row = ws.row(row_idx)
                    try:
                        lat = float(row[lat_col].value) if lat_col is not None and row[lat_col].value else None
                        lon = float(row[lon_col].value) if lon_col is not None and row[lon_col].value else None
                        name = str(row[name_col].value) if name_col is not None and row[name_col].value else f"Toplanma Alanı ({excel_file.stem})"
                        
                        if lat and lon:
                            features.append({
                                "type": "Feature",
                                "geometry": {"type": "Point", "coordinates": [lon, lat]},
                                "properties": {"name": name, "ilçe": excel_file.stem.split("_")[0]}
                            })
                    except (ValueError, TypeError, IndexError):
                        continue
        
        except Exception as e:
            print(f"Excel okuma hatası {excel_file}: {e}")
            continue
    
    return {"type": "FeatureCollection", "features": features}

# ============================
# FIRMS cache (background)
# ============================

DATA_DIR = Path("data")
DATA_DIR.mkdir(parents=True, exist_ok=True)
FIRMS_CACHE_PATH = DATA_DIR / "firms.json"

# Optional: İzmir administrative boundary (GeoJSON) for clipping
IZMIR_GEO_PATH = Path("static/data/izmir.geojson")
IZMIR_POLYS: list | None = None  # list of polygons; each polygon: list of linear rings [[(lon,lat),...], holes...]
IZMIR_BBOX_DEFAULT = (26.230389, 37.818402, 28.495245, 39.392935)

def _load_izmir_polys():
    global IZMIR_POLYS
    try:
        if not IZMIR_GEO_PATH.exists():
            IZMIR_POLYS = None
            return
        with IZMIR_GEO_PATH.open("r", encoding="utf-8") as f:
            gj = json.load(f)
        def collect_geom(geom):
            polys = []
            if not geom:
                return polys
            gtype = geom.get("type")
            coords = geom.get("coordinates")
            if gtype == "Polygon":
                # coords: [ring0, ring1(hole), ...]
                rings = []
                for ring in coords or []:
                    rings.append([(float(x), float(y)) for x, y in ring])
                if rings:
                    polys.append(rings)
            elif gtype == "MultiPolygon":
                for poly in coords or []:
                    rings = []
                    for ring in poly or []:
                        rings.append([(float(x), float(y)) for x, y in ring])
                    if rings:
                        polys.append(rings)
            return polys
        polys = []
        if gj.get("type") == "FeatureCollection":
            for feat in gj.get("features", []):
                polys.extend(collect_geom(feat.get("geometry")))
        elif gj.get("type") == "Feature":
            polys.extend(collect_geom(gj.get("geometry")))
        else:
            polys.extend(collect_geom(gj))
        IZMIR_POLYS = polys or None
    except Exception:
        IZMIR_POLYS = None

def _point_in_ring(lon: float, lat: float, ring: list[tuple[float, float]]) -> bool:
    inside = False
    n = len(ring)
    for i in range(n):
        x1, y1 = ring[i]
        x2, y2 = ring[(i + 1) % n]
        # Ray casting, considering (lat) as y, (lon) as x
        if ((y1 > lat) != (y2 > lat)):
            x_at_y = (x2 - x1) * (lat - y1) / (y2 - y1 + 1e-15) + x1
            if lon < x_at_y:
                inside = not inside
    return inside

def _is_in_izmir(lat: float, lon: float) -> bool:
    # Try polygon first
    if IZMIR_POLYS is None:
        _load_izmir_polys()
    if IZMIR_POLYS:
        for rings in IZMIR_POLYS:
            outer = rings[0]
            holes = rings[1:] if len(rings) > 1 else []
            if _point_in_ring(lon, lat, outer):
                for hole in holes:
                    if _point_in_ring(lon, lat, hole):
                        return False
                return True
        # not in any polygon
        return False
    # Fallback to bbox
    minx, miny, maxx, maxy = IZMIR_BBOX_DEFAULT
    return (minx <= lon <= maxx) and (miny <= lat <= maxy)

# ============================
# Admin: import shelters (Toplanma Alanları)
# ============================
STATIC_DATA_DIR = Path("static/data")
STATIC_DATA_DIR.mkdir(parents=True, exist_ok=True)
SHELTERS_PATH = STATIC_DATA_DIR / "assembly_points.geojson"
SHELTERS_DIR = STATIC_DATA_DIR / "toplanma-alanları-izmir"

def _save_geojson_to_file(obj, dest: Path):
    tmp = dest.with_suffix(".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(obj, f)
    tmp.replace(dest)

def _download_json(url: str, timeout: float = 20.0):
    import requests
    r = requests.get(url, timeout=timeout)
    r.raise_for_status()
    return r.json()

def _arcgis_to_geojson(service_url: str):
    # Normalize ArcGIS Feature/Map Server URL to a layer query URL returning GeoJSON
    url = service_url.strip()
    # If URL already includes /query, assume it's good; ensure f=geojson
    if re.search(r"/query\b", url):
        sep = '&' if '?' in url else '?'
        return f"{url}{sep}f=geojson"
    # Detect layer id
    m = re.search(r"/(FeatureServer|MapServer)(/\d+)?/?$", url)
    if m:
        layer_part = m.group(2) or "/0"
        base = url.rstrip('/') + layer_part + "/query"
    else:
        # If URL already ends with /<layer>, add /query
        if re.search(r"/\d+/?$", url):
            base = url.rstrip('/') + "/query"
        else:
            # Best-effort default: append /0/query
            base = url.rstrip('/') + "/0/query"
    params = "where=1=1&outFields=*&f=geojson&outSR=4326&resultRecordCount=20000"
    return base + ("?" + params)

@app.get("/api/admin/import_shelters")
def import_shelters(url: str = Query(..., description="GeoJSON or ArcGIS Feature/MapServer URL"),
                    type: str | None = Query(None, description="geojson|arcgis (auto if omitted)")):
    """
    Downloads shelter (assembly points) data from a given URL and stores it to static/data/assembly_points.geojson.
    - Accepts direct GeoJSON URLs or ArcGIS FeatureServer/MapServer endpoints (auto-converted to GeoJSON query).
    """
    try:
        src = url
        kind = (type or '').lower().strip()
        if not kind:
            if re.search(r"FeatureServer|MapServer", src):
                kind = "arcgis"
            else:
                kind = "geojson"
        if kind == "arcgis":
            src = _arcgis_to_geojson(src)
        data = _download_json(src)
        # Basic validation
        if not isinstance(data, dict) or data.get("type") not in ("FeatureCollection", "Feature"):
            return JSONResponse({"error": "invalid_geojson", "note": "Response is not a GeoJSON Feature/FeatureCollection", "source": src}, status_code=400)
        _save_geojson_to_file(data, SHELTERS_PATH)
        return {"ok": True, "saved": str(SHELTERS_PATH), "source": url, "resolved": src, "features": len((data.get("features") or []))}
    except Exception as e:
        return JSONResponse({"error": f"import_failed: {e}"}, status_code=500)

@app.get("/api/shelters_manifest")
def shelters_manifest():
    """List locally available shelter files in static/data/toplanma-alanları-izmir.
    Returns URLs suitable for fetching from the browser."""
    files = []
    if SFT := SHELTERS_PATH.exists():
        files.append("/static/data/assembly_points.geojson")
    if SFT:
        pass
    if SFT or SHELTERS_DIR.exists():
        # Enumerate known formats in subdir
        if SHELTERS_DIR.exists():
            for p in sorted(SHELTERS_DIR.iterdir()):
                if not p.is_file():
                    continue
                ext = p.suffix.lower()
                if ext in (".xlsx", ".xls", ".csv", ".geojson", ".json"):
                    files.append("/static/data/toplanma-alanları-izmir/" + p.name)
    return {"files": files}

def _read_json(path: Path):
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

def _write_json(path: Path, obj):
    tmp = path.with_suffix(".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(obj, f)
    tmp.replace(path)

async def _poll_firms_forever():
    if fetch_firms_geojson is None:
        return
    poll_minutes = float(os.getenv("FIRMS_POLL_MINUTES", "15"))
    interval = max(3.0, poll_minutes) * 60.0
    while True:
        try:
            data = fetch_firms_geojson(day_range=3)
            now_iso = time.strftime("%Y-%m-%dT%H:%M:%S")
            if isinstance(data, dict) and data.get("type") == "FeatureCollection":
                payload = {"updated_at": now_iso, "features": data.get("features", [])}
                _write_json(FIRMS_CACHE_PATH, payload)
            else:
                # preserve previous cache; write error status for visibility
                payload = {"updated_at": now_iso, "error": data}
                _write_json(FIRMS_CACHE_PATH, payload)
        except Exception as e:
            payload = {"updated_at": time.strftime("%Y-%m-%dT%H:%M:%S"), "error": f"poll_error: {e}"}
            _write_json(FIRMS_CACHE_PATH, payload)
        await asyncio.sleep(interval)

@app.on_event("startup")
async def _startup():
    # start background polling (fire-and-forget)
    try:
        asyncio.get_event_loop().create_task(_poll_firms_forever())
    except Exception:
        pass

@app.get("/api/fires_cached")
async def fires_cached(refresh: bool = Query(False), day_range: int = Query(3)):
    """Serve cached FIRMS. Optionally trigger a refresh fetch."""
    if refresh and fetch_firms_geojson is not None:
        try:
            data = fetch_firms_geojson(day_range=day_range)
            now_iso = time.strftime("%Y-%m-%dT%H:%M:%S")
            if isinstance(data, dict) and data.get("type") == "FeatureCollection":
                payload = {"updated_at": now_iso, "features": data.get("features", [])}
            else:
                payload = {"updated_at": now_iso, "error": data}
            _write_json(FIRMS_CACHE_PATH, payload)
        except Exception as e:
            return JSONResponse({"error": f"refresh_failed: {e}"}, status_code=500)

    cached = _read_json(FIRMS_CACHE_PATH) or {}
    if not cached:
        # If cache empty, try an immediate best-effort fetch once
        if fetch_firms_geojson is not None:
            try:
                data = fetch_firms_geojson(day_range=day_range)
                now_iso = time.strftime("%Y-%m-%dT%H:%M:%S")
                if isinstance(data, dict) and data.get("type") == "FeatureCollection":
                    cached = {"updated_at": now_iso, "features": data.get("features", [])}
                    _write_json(FIRMS_CACHE_PATH, cached)
            except Exception:
                pass
    # format as standard FeatureCollection with meta
    features = cached.get("features", [])
    meta = {k: v for k, v in cached.items() if k != "features"}
    return JSONResponse({"type": "FeatureCollection", "features": features, "meta": meta})

# ============================
# Risk stub (grid-based, proximity to fires)
# ============================

def _parse_bbox_str(bbox_str: str):
    try:
        minx, miny, maxx, maxy = [float(x) for x in bbox_str.split(",")]
        return minx, miny, maxx, maxy
    except Exception:
        # Fallback to İzmir approx
        return 26.230389, 37.818402, 28.495245, 39.392935

def _haversine_km(lat1, lon1, lat2, lon2):
    from math import radians, sin, cos, asin, sqrt
    R = 6371.0
    dlat = radians(lat2 - lat1)
    dlon = radians(lon2 - lon1)
    a = sin(dlat/2)**2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a))
    return R * c

def _bearing_deg(lat1, lon1, lat2, lon2):
    # Bearing from (lat1,lon1) to (lat2,lon2), degrees clockwise from North
    lat1r, lat2r = math.radians(lat1), math.radians(lat2)
    dlon = math.radians(lon2 - lon1)
    x = math.sin(dlon) * math.cos(lat2r)
    y = math.cos(lat1r) * math.sin(lat2r) - math.sin(lat1r) * math.cos(lat2r) * math.cos(dlon)
    br = math.degrees(math.atan2(x, y))
    return (br + 360.0) % 360.0

def _angle_diff_deg(a, b):
    d = abs((a - b + 180) % 360 - 180)
    return d

def _bbox_intersection(a, b):
    ax1, ay1, ax2, ay2 = a
    bx1, by1, bx2, by2 = b
    ix1 = max(ax1, bx1)
    iy1 = max(ay1, by1)
    ix2 = min(ax2, bx2)
    iy2 = min(ay2, by2)
    if ix2 <= ix1 or iy2 <= iy1:
        return None
    return ix1, iy1, ix2, iy2

@app.get("/api/risk")
def risk(cell_km: float = Query(8.0),
         bbox: str | None = Query(None),
         use_default_wind: bool = Query(False)):
    """
    Wind-aligned risk over a bbox (default İzmir).
    - Risk decays with distance from fires, slower downwind, faster upwind.
    - Cells are rotated rectangles aligned to downwind direction.
    Returns FeatureCollection with `risk` in [0,1].
    """
    try:
        try:
            from app.firms import IZMIR_BBOX
            izmir_bbox = _parse_bbox_str(IZMIR_BBOX)
        except Exception:
            izmir_bbox = (26.230389, 37.818402, 28.495245, 39.392935)
        req_bbox = _parse_bbox_str(bbox) if bbox else izmir_bbox
        inter = _bbox_intersection(req_bbox, izmir_bbox)
        if not inter:
            return JSONResponse({"error": "bbox_outside_izmir"}, status_code=400)
        minx, miny, maxx, maxy = inter

        cached = _read_json(FIRMS_CACHE_PATH) or {}
        fire_pts = []
        for f in cached.get("features", []):
            try:
                lon, lat = f.get("geometry", {}).get("coordinates", [None, None])
                if lat is not None and lon is not None:
                    fire_pts.append((lat, lon))
            except Exception:
                continue

        # Build grid: approx deg step from km (lon adjusts by cos(lat))
        mid_lat = (miny + maxy) / 2.0
        mid_lon = (minx + maxx) / 2.0
        # Get wind for bbox center (optional default to avoid network stall)
        if use_default_wind:
            wind_from, wind_ms = 240.0, 6.0
        else:
            try:
                w = get_wind(mid_lat, mid_lon)
                wind_from = float(w.get("deg", 240.0))
                wind_ms = float(w.get("speed_ms", 6.0))
            except Exception:
                wind_from, wind_ms = 240.0, 6.0
        wind_down = (wind_from + 180.0) % 360.0

        deg_per_km_lat = 1.0 / 111.32
        deg_per_km_lon = deg_per_km_lat / max(0.1, abs(math.cos(math.radians(mid_lat))))
        # Base sampling step uses shorter crosswind size for more detail
        cross_km = max(2.0, min(50.0, cell_km))
        along_km = cross_km * 1.8  # elongated along downwind for visual clarity
        dlat = cross_km * deg_per_km_lat
        dlon = cross_km * deg_per_km_lon

        # Cap cell count to keep rendering responsive
        lat_span = maxy - miny
        lon_span = maxx - minx
        n_lat = max(1, int(lat_span / max(1e-9, dlat)))
        n_lon = max(1, int(lon_span / max(1e-9, dlon)))
        max_cells = 800
        total = n_lat * n_lon
        if total > max_cells:
            scale = (total / max_cells) ** 0.5
            cross_km *= scale
            along_km = cross_km * 1.8
            dlat = cross_km * deg_per_km_lat
            dlon = cross_km * deg_per_km_lon

        features = []
        lat = miny
        while lat < maxy:
            lon = minx
            while lon < maxx:
                lat_c = lat + dlat/2
                lon_c = lon + dlon/2
                # distance to nearest fire (anisotropic downwind)
                if fire_pts:
                    d_eff_min = None
                    for fy, fx in fire_pts:
                        d_km = _haversine_km(fy, fx, lat_c, lon_c)
                        if d_km == 0:
                            d_eff = 0.0
                        else:
                            br = _bearing_deg(fy, fx, lat_c, lon_c)
                            delta = _angle_diff_deg(br, wind_down)
                            # directional weight in [0,1]
                            wdir = 0.5 + 0.5 * math.cos(math.radians(delta))
                            beta = 0.6
                            d_eff = d_km * (1.0 - beta * wdir)
                            if delta > 150:
                                d_eff *= 1.2
                        d_eff_min = d_eff if d_eff_min is None else min(d_eff_min, d_eff)
                    L = 8.0 + 0.8 * wind_ms
                    risk = float(max(0.0, min(1.0, math.exp(-(d_eff_min or 0.0) / L))))
                else:
                    risk = 0.15

                # Build wind-aligned rotated rectangle polygon
                half_u = (along_km * 1000.0) / 2.0
                half_v = (cross_km * 1000.0) / 2.0
                theta = math.radians(wind_down)
                ux, uy = math.sin(theta), math.cos(theta)
                vx, vy = math.sin(theta + math.pi/2), math.cos(theta + math.pi/2)
                corners = []
                for su in (-1, 1):
                    for sv in (-1, 1):
                        dx = su * half_u * ux + sv * half_v * vx
                        dy = su * half_u * uy + sv * half_v * vy
                        dlat_m, dlon_m = meters_to_deg(lat_c, dx, dy)
                        corners.append([lon_c + dlon_m, lat_c + dlat_m])
                poly = [corners[0], corners[1], corners[3], corners[2], corners[0]]
                features.append({
                    "type": "Feature",
                    "geometry": {"type": "Polygon", "coordinates": [poly]},
                    "properties": {
                        "risk": round(risk, 3),
                        "cell_km": round(cross_km, 2),
                        "wind_from": round(wind_from, 1),
                        "wind_speed_ms": round(wind_ms, 1)
                    }
                })
                lon += dlon
            lat += dlat

        return JSONResponse({"type": "FeatureCollection", "features": features, "meta": {"updated_at": (cached or {}).get("updated_at")}})
    except Exception as e:
        # Return error details to help diagnose
        return JSONResponse({"error": f"risk_failed: {e}"}, status_code=500)

# Helpers to compute scalar risk for a point
def _compute_wind(lat: float, lon: float, use_default: bool = False):
    if use_default:
        return 240.0, 6.0
    try:
        w = get_wind(lat, lon)
        return float(w.get("deg", 240.0)), float(w.get("speed_ms", 6.0))
    except Exception:
        return 240.0, 6.0

def _risk_value(lat: float, lon: float, fire_pts, wind_down: float, wind_ms: float) -> float:
    if not fire_pts:
        return 0.15
    d_eff_min = None
    for fy, fx in fire_pts:
        d_km = _haversine_km(fy, fx, lat, lon)
        if d_km == 0:
            d_eff = 0.0
        else:
            br = _bearing_deg(fy, fx, lat, lon)
            delta = _angle_diff_deg(br, wind_down)
            wdir = 0.5 + 0.5 * math.cos(math.radians(delta))
            beta = 0.6
            d_eff = d_km * (1.0 - beta * wdir)
            if delta > 150:
                d_eff *= 1.2
        d_eff_min = d_eff if d_eff_min is None else min(d_eff_min, d_eff)
    L = 8.0 + 0.8 * wind_ms
    return float(max(0.0, min(1.0, math.exp(-(d_eff_min or 0.0) / L))))

@app.get("/api/risk_grid")
def risk_grid(cell_km: float = Query(6.0),
              bbox: str | None = Query(None),
              use_default_wind: bool = Query(True),
              max_points: int = Query(1500)):
    """
    Returns a scalar risk grid over bbox for contour rendering.
    - Output: { type: "RiskGrid", bbox, nrows, ncols, dlat, dlon, values (row-major), meta }
    """
    try:
        try:
            from app.firms import IZMIR_BBOX
            izmir_bbox = _parse_bbox_str(IZMIR_BBOX)
        except Exception:
            izmir_bbox = (26.230389, 37.818402, 28.495245, 39.392935)
        req_bbox = _parse_bbox_str(bbox) if bbox else izmir_bbox
        inter = _bbox_intersection(req_bbox, izmir_bbox)
        if not inter:
            return JSONResponse({"error": "bbox_outside_izmir"}, status_code=400)
        minx, miny, maxx, maxy = inter

        cached = _read_json(FIRMS_CACHE_PATH) or {}
        fire_pts = []
        for f in cached.get("features", []):
            try:
                lon, lat = f.get("geometry", {}).get("coordinates", [None, None])
                if lat is not None and lon is not None:
                    fire_pts.append((lat, lon))
            except Exception:
                continue

        mid_lat = (miny + maxy) / 2.0
        mid_lon = (minx + maxx) / 2.0
        wind_from, wind_ms = _compute_wind(mid_lat, mid_lon, use_default=use_default_wind)
        wind_down = (wind_from + 180.0) % 360.0

        # Derive grid steps
        deg_per_km_lat = 1.0 / 111.32
        deg_per_km_lon = deg_per_km_lat / max(0.1, abs(math.cos(math.radians(mid_lat))))
        step_lat = max(0.001, cell_km * deg_per_km_lat)
        step_lon = max(0.001, cell_km * deg_per_km_lon)

        lat_span = maxy - miny
        lon_span = maxx - minx
        nrows = max(2, int(lat_span / step_lat) + 1)
        ncols = max(2, int(lon_span / step_lon) + 1)

        # Cap total points
        total = nrows * ncols
        if total > max_points:
            scale = (total / max_points) ** 0.5
            step_lat *= scale
            step_lon *= scale
            nrows = max(2, int(lat_span / step_lat) + 1)
            ncols = max(2, int(lon_span / step_lon) + 1)

        # Recompute exact step to land on max
        if nrows > 1:
            step_lat = lat_span / (nrows - 1)
        if ncols > 1:
            step_lon = lon_span / (ncols - 1)

        values = []  # row-major
        for r in range(nrows):
            lat = miny + r * step_lat
            for c in range(ncols):
                lon = minx + c * step_lon
                if not _is_in_izmir(lat, lon):
                    rv = 0.0
                else:
                    rv = _risk_value(lat, lon, fire_pts, wind_down, wind_ms)
                values.append(round(rv, 4))

        return JSONResponse({
            "type": "RiskGrid",
            "bbox": [minx, miny, maxx, maxy],
            "nrows": nrows,
            "ncols": ncols,
            "dlat": step_lat,
            "dlon": step_lon,
            "values": values,
            "meta": {
                "wind_from": round(wind_from, 1),
                "wind_speed_ms": round(wind_ms, 1),
                "updated_at": (cached or {}).get("updated_at")
            }
        })
    except Exception as e:
        return JSONResponse({"error": f"risk_grid_failed: {e}"}, status_code=500)

@app.get("/api/risk_test")
def risk_test(bbox: str | None = Query(None)):
    """Very small fixed grid to validate rendering quickly."""
    try:
        try:
            from app.firms import IZMIR_BBOX
            bbox_str = bbox or IZMIR_BBOX
        except Exception:
            bbox_str = bbox or "26.230389,37.818402,28.495245,39.392935"
        minx, miny, maxx, maxy = _parse_bbox_str(bbox_str)
        cx = (minx + maxx) / 2.0
        cy = (miny + maxy) / 2.0
        step_lon = (maxx - minx) / 40.0 or 0.02
        step_lat = (maxy - miny) / 40.0 or 0.02
        features = []
        for i in range(-2, 3):
            for j in range(-2, 3):
                lon0 = cx + i * step_lon
                lat0 = cy + j * step_lat
                lon1 = lon0 + step_lon
                lat1 = lat0 + step_lat
                risk = max(0.0, 1.0 - (abs(i) + abs(j)) * 0.2)
                poly = [[lon0, lat0],[lon1, lat0],[lon1, lat1],[lon0, lat1],[lon0, lat0]]
                features.append({
                    "type": "Feature",
                    "geometry": {"type": "Polygon", "coordinates": [poly]},
                    "properties": {"risk": round(risk,3), "test": True}
                })
        return JSONResponse({"type": "FeatureCollection", "features": features, "meta": {"test": True}})
    except Exception as e:
        return JSONResponse({"error": f"risk_test_failed: {e}"}, status_code=500)
